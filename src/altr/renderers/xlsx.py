from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .. import theme
from ..schemas import Sheet, SpreadsheetSpec

_CHART_CLASSES = {"bar": BarChart, "line": LineChart, "pie": PieChart}
_CHART_ROWS = 15  # vertical space to leave per chart when stacking

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", start_color=theme.HEADER_FILL)
_BAND_FILL = PatternFill("solid", start_color=theme.ROW_BAND)
_TOTAL_FONT = Font(bold=True, color=theme.INK)
_TOTAL_BORDER = Border(top=Side(style="thin", color="C3C2B7"))

_MIN_COL_WIDTH = 9
_MAX_COL_WIDTH = 40


def render_spreadsheet(spec: SpreadsheetSpec, out_dir: Path) -> Path:
    from . import output_path

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in spec.sheets:
        write_sheet(wb, sheet)

    path = output_path(out_dir, spec.filename, ".xlsx")
    wb.save(str(path))
    return path


def write_sheet(wb: Workbook, sheet: Sheet) -> None:
    """Create and fill one styled worksheet (used by render and edit)."""
    ws = wb.create_sheet(title=sheet.name)
    width = max((len(row) for row in sheet.rows), default=len(sheet.columns))
    first_data_row = 1
    if sheet.columns:
        for i, col in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=1, column=i, value=col.header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        if sheet.freeze_header:
            ws.freeze_panes = "A2"
        first_data_row = 2

    total_row = _total_row_index(sheet.rows)
    for r, row in enumerate(sheet.rows):
        is_total = r == total_row
        for c in range(1, width + 1):
            cell = ws.cell(
                row=r + first_data_row,
                column=c,
                value=row[c - 1] if c <= len(row) else None,
            )
            if is_total:
                cell.font = _TOTAL_FONT
                cell.border = _TOTAL_BORDER
            elif r % 2 == 1:
                cell.fill = _BAND_FILL

    _size_columns(ws, sheet, width)
    _add_charts(ws, sheet, first_data_row)


def _total_row_index(rows: list[list]) -> int | None:
    """The last row, if it reads as a summary row (e.g. 'Total', 'Grand total')."""
    if rows and rows[-1] and isinstance(rows[-1][0], str) and "total" in rows[-1][0].lower():
        return len(rows) - 1
    return None


def _size_columns(ws: Worksheet, sheet: Sheet, width: int) -> None:
    """Fit column widths to content unless the model specified one."""
    for c in range(1, width + 1):
        letter = get_column_letter(c)
        if c <= len(sheet.columns) and sheet.columns[c - 1].width:
            ws.column_dimensions[letter].width = sheet.columns[c - 1].width
            continue
        lengths = [len(sheet.columns[c - 1].header)] if c <= len(sheet.columns) else []
        lengths += [
            len(str(row[c - 1]))
            for row in sheet.rows
            if len(row) >= c and row[c - 1] is not None and not _is_formula(row[c - 1])
        ]
        if lengths:
            fitted = max(lengths) + 3  # padding; headers are bold and wider
            ws.column_dimensions[letter].width = max(
                _MIN_COL_WIDTH, min(_MAX_COL_WIDTH, fitted)
            )


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _add_charts(ws: Worksheet, sheet: Sheet, first_data_row: int) -> None:
    if not sheet.charts or not sheet.rows:
        return
    # Plot data rows only - a trailing 'Total' row would dwarf the categories.
    data_rows = len(sheet.rows) - (1 if _total_row_index(sheet.rows) is not None else 0)
    if not data_rows:
        return
    last_row = first_data_row + data_rows - 1
    data_cols = max(len(row) for row in sheet.rows)
    anchor_col = get_column_letter(data_cols + 2)

    for i, chart_spec in enumerate(sheet.charts):
        value_columns = chart_spec.value_columns or _numeric_columns(
            sheet.rows, exclude=chart_spec.label_column
        )
        if not value_columns:
            continue
        chart = _CHART_CLASSES[chart_spec.kind]()
        chart.title = chart_spec.title
        chart.width = 16
        chart.height = 9
        has_header = bool(sheet.columns)
        for col in value_columns:
            data = Reference(
                ws,
                min_col=col,
                min_row=1 if has_header else first_data_row,
                max_row=last_row,
            )
            chart.add_data(data, titles_from_data=has_header)
        categories = Reference(
            ws, min_col=chart_spec.label_column, min_row=first_data_row, max_row=last_row
        )
        chart.set_categories(categories)
        _style_chart(chart, chart_spec.kind, len(value_columns))
        ws.add_chart(chart, f"{anchor_col}{2 + i * _CHART_ROWS}")


def _style_chart(chart, kind: str, series_count: int) -> None:
    if kind == "pie":
        return  # Excel varies slice colors; the legend names the slices
    if series_count == 1:
        chart.legend = None  # a single series is named by the chart title
    for i, series in enumerate(chart.series):
        color = theme.SERIES[i % len(theme.SERIES)]
        if kind == "line":
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 28575  # 2.25pt in EMU
            series.smooth = False
        else:
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.noFill = True
    # Explicit axis/gridline styling: unstyled openpyxl axes render with harsh
    # black gridlines (and hidden labels in some viewers).
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    hairline = LineProperties(solidFill=theme.GRIDLINE, w=9525)  # 0.75pt
    chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=hairline))
    chart.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="C3C2B7"))


def _numeric_columns(rows: list[list], exclude: int) -> list[int]:
    """1-based indexes of columns holding any numeric data, used when the
    model omitted value_columns."""
    width = max(len(row) for row in rows)
    return [
        c
        for c in range(1, width + 1)
        if c != exclude
        and any(
            len(row) >= c and isinstance(row[c - 1], (int, float)) and not isinstance(row[c - 1], bool)
            for row in rows
        )
    ]
