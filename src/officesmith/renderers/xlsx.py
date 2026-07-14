from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..schemas import SpreadsheetSpec


def render_spreadsheet(spec: SpreadsheetSpec, out_dir: Path) -> Path:
    from . import output_path

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in spec.sheets:
        ws = wb.create_sheet(title=sheet.name)
        first_data_row = 1
        if sheet.columns:
            for i, col in enumerate(sheet.columns, start=1):
                cell = ws.cell(row=1, column=i, value=col.header)
                cell.font = Font(bold=True)
                if col.width:
                    ws.column_dimensions[get_column_letter(i)].width = col.width
            if sheet.freeze_header:
                ws.freeze_panes = "A2"
            first_data_row = 2

        for r, row in enumerate(sheet.rows, start=first_data_row):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

    path = output_path(out_dir, spec.filename, ".xlsx")
    wb.save(str(path))
    return path
